# ============================================================
#  Archivo: backend/routers/reportes.py
# ============================================================
from fastapi import APIRouter, HTTPException
from backend.db import get_connection
from datetime import datetime, timedelta

router = APIRouter()

def fechas_default():
    fin = datetime.now().strftime("%Y-%m-%d")
    ini = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    return ini, fin

@router.get("/volumen")
def reporte_volumen(fecha_inicio: str = None, fecha_fin: str = None):
    if not fecha_inicio or not fecha_fin:
        fecha_inicio, fecha_fin = fechas_default()
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            WITH VentasHistoricas AS (
                SELECT dv.id_producto, p.nombre, SUM(dv.cantidad) as volumen_total
                FROM detalle_ventas dv
                JOIN ventas v ON dv.id_venta = v.id_venta
                JOIN productos p ON dv.id_producto = p.id_producto
                WHERE v.fecha::date BETWEEN %s AND %s
                GROUP BY dv.id_producto, p.nombre
            )
            SELECT id_producto, nombre, volumen_total,
                   RANK() OVER (ORDER BY volumen_total DESC) as ranking
            FROM VentasHistoricas
            ORDER BY ranking ASC
        """, (fecha_inicio, fecha_fin))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"id": r[0], "producto": r[1], "volumen": float(r[2]), "ranking": r[3]} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/rentabilidad")
def reporte_rentabilidad(fecha_inicio: str = None, fecha_fin: str = None):
    if not fecha_inicio or not fecha_fin:
        fecha_inicio, fecha_fin = fechas_default()
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            WITH CostoPromedio AS (
                SELECT id_producto, AVG(costo) as costo_medio
                FROM produccion
                WHERE fecha_produccion BETWEEN %s AND %s
                GROUP BY id_producto
            )
            SELECT dv.id_producto, p.nombre,
                   SUM(dv.total_fila) as ingreso_bruto,
                   SUM(dv.cantidad * COALESCE(cp.costo_medio, 0)) as costo_manufactura,
                   SUM(dv.total_fila) - SUM(dv.cantidad * COALESCE(cp.costo_medio, 0)) as ganancia_neta
            FROM detalle_ventas dv
            JOIN ventas v ON dv.id_venta = v.id_venta
            JOIN productos p ON dv.id_producto = p.id_producto
            LEFT JOIN CostoPromedio cp ON p.id_producto = cp.id_producto
            WHERE v.fecha::date BETWEEN %s AND %s
            GROUP BY dv.id_producto, p.nombre
            ORDER BY ganancia_neta DESC LIMIT 5
        """, (fecha_inicio, fecha_fin, fecha_inicio, fecha_fin))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"id": r[0], "producto": r[1], "ingreso_bruto": float(r[2]),
                 "costo_manufactura": float(r[3]), "ganancia_neta": float(r[4])} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/mermas")
def reporte_mermas():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT p.nombre,
                   SUM(pe.cantidad_estimada) as proyectado,
                   SUM(pe.cantidad_lograda) as real,
                   SUM(pe.cantidad_estimada - pe.cantidad_lograda) as merma,
                   ROUND(
                       (SUM(pe.cantidad_estimada - pe.cantidad_lograda) * 100.0) /
                       NULLIF(SUM(pe.cantidad_estimada), 0), 2
                   ) as pct_merma
            FROM proceso_elaboracion pe
            JOIN productos p ON pe.id_producto = p.id_producto
            GROUP BY p.nombre
            ORDER BY pct_merma DESC NULLS LAST
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"producto": r[0], "proyectado": r[1], "real": r[2],
                 "merma": r[3], "pct_merma": float(r[4]) if r[4] else 0.0} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/ventas-semana")
def ventas_semana():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT
                DATE(fecha) as dia,
                COUNT(*) as num_ventas,
                COALESCE(SUM(total_neto), 0) as ingresos
            FROM ventas
            WHERE fecha >= CURRENT_DATE - INTERVAL '6 days'
            GROUP BY DATE(fecha)
            ORDER BY dia ASC
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        # Rellenar días sin ventas con 0
        from datetime import date, timedelta
        resultado = {}
        for i in range(7):
            d = (date.today() - timedelta(days=6-i)).isoformat()
            resultado[d] = {"dia": d, "ventas": 0, "ingresos": 0.0}
        for r in rows:
            k = str(r[0])
            if k in resultado:
                resultado[k] = {"dia": k, "ventas": r[1], "ingresos": float(r[2])}
        return list(resultado.values())
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/dashboard-kpis")
def dashboard_kpis():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM ventas WHERE fecha::date = CURRENT_DATE")
        ventas_hoy = cur.fetchone()[0]
        cur.execute("SELECT COALESCE(SUM(total_neto), 0) FROM ventas WHERE fecha::date = CURRENT_DATE")
        ingresos_hoy = float(cur.fetchone()[0])
        cur.execute("SELECT COUNT(*) FROM inventario WHERE stock_actual < 10")
        stock_bajo = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM empleados WHERE activo = 1")
        empleados_activos = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM produccion WHERE fecha_produccion = CURRENT_DATE")
        lotes_hoy = cur.fetchone()[0]
        cur.close(); conn.close()
        return {
            "ventas_hoy": ventas_hoy,
            "ingresos_hoy": ingresos_hoy,
            "stock_bajo": stock_bajo,
            "empleados_activos": empleados_activos,
            "lotes_hoy": lotes_hoy
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Exportar Reporte XLSX ─────────────────────────────────────
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

def estilo_encabezado(cell, color="D97706"):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    cell.fill      = PatternFill("solid", start_color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def estilo_celda(cell, bold=False):
    cell.font      = Font(name="Arial", size=10, bold=bold)
    cell.alignment = Alignment(vertical="center")
    thin = Side(style="thin", color="EEEEEE")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

@router.get("/exportar-xlsx")
def exportar_xlsx(tipo: str, fecha_inicio: str = None, fecha_fin: str = None):
    conn = get_connection()
    cur  = conn.cursor()
    wb   = Workbook()
    ws   = wb.active

    hoy = datetime.today().strftime("%d/%m/%Y")

    if tipo == "volumen":
        ws.title = "Volumen de Ventas"
        fi = fecha_inicio or "2000-01-01"
        ff = fecha_fin   or "2099-12-31"
        cur.execute("""
            SELECT p.nombre, SUM(dv.cantidad) as unidades,
                   SUM(dv.total_fila) as ingresos,
                   ROUND(SUM(dv.total_fila)/NULLIF(SUM(dv.cantidad),0),2) as precio_prom
            FROM detalle_ventas dv
            JOIN productos p ON dv.id_producto = p.id_producto
            JOIN ventas v ON dv.id_venta = v.id_venta
            WHERE v.fecha BETWEEN %s AND %s
            GROUP BY p.nombre ORDER BY unidades DESC
        """, (fi, ff))
        rows = cur.fetchall()

        # Título
        ws.merge_cells("A1:D1")
        ws["A1"] = f"Reporte de Volumen de Ventas — {fi} al {ff}"
        ws["A1"].font = Font(bold=True, size=14, name="Arial", color="3D1A00")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:D2")
        ws["A2"] = f"Generado el {hoy}"
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].font = Font(size=9, color="888888", name="Arial")

        # Encabezados
        headers = ["Producto", "Unidades Vendidas", "Ingresos ($)", "Precio Promedio ($)"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            estilo_encabezado(c)

        # Datos
        for i, r in enumerate(rows, 5):
            ws.cell(row=i, column=1, value=r[0])
            ws.cell(row=i, column=2, value=int(r[1]))
            ws.cell(row=i, column=3, value=float(r[2]))
            ws.cell(row=i, column=4, value=float(r[3]) if r[3] else 0)
            for col in range(1, 5):
                c = ws.cell(row=i, column=col)
                estilo_celda(c)
                if i % 2 == 0:
                    c.fill = PatternFill("solid", start_color="FFF8EC")

        # Totales
        last = len(rows) + 5
        ws.cell(row=last, column=1, value="TOTAL")
        ws.cell(row=last, column=2, value=f"=SUM(B5:B{last-1})")
        ws.cell(row=last, column=3, value=f"=SUM(C5:C{last-1})")
        ws.cell(row=last, column=4, value="")
        for col in range(1, 5):
            estilo_encabezado(ws.cell(row=last, column=col), color="3D1A00")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 22

    elif tipo == "rentabilidad":
        ws.title = "Rentabilidad"
        fi = fecha_inicio or "2000-01-01"
        ff = fecha_fin   or "2099-12-31"
        cur.execute("""
            SELECT p.nombre,
                   SUM(dv.cantidad) as unidades,
                   SUM(dv.total_fila) as ingresos,
                   ROUND(AVG(prod.costo),4) as costo_unit,
                   SUM(dv.total_fila) - SUM(dv.cantidad * prod.costo) as ganancia
            FROM detalle_ventas dv
            JOIN productos p ON dv.id_producto = p.id_producto
            JOIN ventas v ON dv.id_venta = v.id_venta
            LEFT JOIN produccion prod ON prod.id_producto = p.id_producto
            WHERE v.fecha BETWEEN %s AND %s
            GROUP BY p.nombre ORDER BY ganancia DESC
        """, (fi, ff))
        rows = cur.fetchall()

        ws.merge_cells("A1:E1")
        ws["A1"] = f"Reporte de Rentabilidad — {fi} al {ff}"
        ws["A1"].font = Font(bold=True, size=14, name="Arial", color="3D1A00")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:E2")
        ws["A2"] = f"Generado el {hoy}"
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].font = Font(size=9, color="888888", name="Arial")

        headers = ["Producto", "Unidades", "Ingresos ($)", "Costo Unit. ($)", "Ganancia ($)"]
        for col, h in enumerate(headers, 1):
            estilo_encabezado(ws.cell(row=4, column=col))
            ws.cell(row=4, column=col).value = h

        for i, r in enumerate(rows, 5):
            vals = [r[0], int(r[1] or 0), float(r[2] or 0), float(r[3] or 0), float(r[4] or 0)]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=i, column=col, value=v)
                estilo_celda(c)
                if i % 2 == 0:
                    c.fill = PatternFill("solid", start_color="FFF8EC")
            # Color ganancia negativa en rojo
            gan = ws.cell(row=i, column=5)
            if float(r[4] or 0) < 0:
                gan.font = Font(color="CC0000", bold=True, name="Arial", size=10)

        last = len(rows) + 5
        ws.cell(row=last, column=1, value="TOTAL")
        ws.cell(row=last, column=2, value=f"=SUM(B5:B{last-1})")
        ws.cell(row=last, column=3, value=f"=SUM(C5:C{last-1})")
        ws.cell(row=last, column=4, value="")
        ws.cell(row=last, column=5, value=f"=SUM(E5:E{last-1})")
        for col in range(1, 6):
            estilo_encabezado(ws.cell(row=last, column=col), color="3D1A00")

        for col, w in enumerate([30, 14, 18, 18, 18], 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    elif tipo == "mermas":
        ws.title = "Mermas Productivas"
        cur.execute("""
            SELECT p.nombre, pe.cantidad_estimada, pe.cantidad_lograda,
                   pe.cantidad_estimada - pe.cantidad_lograda as merma,
                   CASE WHEN pe.cantidad_estimada > 0
                        THEN ROUND(((pe.cantidad_estimada - pe.cantidad_lograda)::numeric / pe.cantidad_estimada)*100, 2)
                        ELSE 0 END as pct_merma,
                   DATE(pe.hora_inicio) as fecha
            FROM proceso_elaboracion pe
            JOIN productos p ON pe.id_producto = p.id_producto
            WHERE pe.cantidad_lograda > 0 AND pe.cantidad_estimada > pe.cantidad_lograda
            ORDER BY pct_merma DESC
        """)
        rows = cur.fetchall()

        ws.merge_cells("A1:F1")
        ws["A1"] = f"Reporte de Mermas Productivas — {hoy}"
        ws["A1"].font = Font(bold=True, size=14, name="Arial", color="3D1A00")
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Producto", "Estimado", "Logrado", "Merma (pzas)", "% Merma", "Fecha"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col, value=h)
            estilo_encabezado(c)

        for i, r in enumerate(rows, 4):
            vals = [r[0], r[1], r[2], r[3], float(r[4]), str(r[5])]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=i, column=col, value=v)
                estilo_celda(c)
                if i % 2 == 0:
                    c.fill = PatternFill("solid", start_color="FFF8EC")
            # % merma en rojo si > 20%
            if float(r[4]) > 20:
                ws.cell(row=i, column=5).font = Font(color="CC0000", bold=True, name="Arial", size=10)

        for col, w in enumerate([28, 12, 12, 16, 12, 14], 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    cur.close(); conn.close()

    # Freeze panes
    ws.freeze_panes = ws.cell(row=5 if tipo != "mermas" else 4, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre_archivo = f"reporte_{tipo}_{datetime.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
    )